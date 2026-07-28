from IPython.display import Audio # Module to play audio in Jupyter notebooks
from numpy import sin, pi, arange
import csv
import openai # OpenAI API for GPT-based language generation
import pandas as pd
from openpyxl import Workbook, load_workbook
import os
# from openai import OpenAI

# Optional: Function to generate a sine wave sound at a specified frequency and duration
def beep(frequency=440, duration=1, sampling_rate=44100):
    t = arange(sampling_rate * duration)
    waveform = sin(2 * pi * frequency * t / sampling_rate)
    return Audio(waveform, rate=sampling_rate, autoplay=True)



openai.api_key = "YOUR_API_KEY_HERE"
openai.api_base = "https://dashscope.aliyuncs.com/compatible-mode/v1"


# # Function to generate ontology competency questions using GPT-4 based on CSV row triples and a given prompt
def generate_questions(rows):
    questions = []
    for row in rows:
        #complete_prompt = f"Subject: {row[0]}, Predicate: {row[1]}, Object: {row[2]}" # Create prompt from CSV row data
        complete_prompt = f"{','.join(row)}?"
        try:
            completion = openai.ChatCompletion.create(
                model="qwen-max",
                messages=[
                    {"role": "system", "content": "As an ontology engineer, Provide competency questions focused on the context provided; avoid using narrative questions. competency questions are the questions that outline the scope of an ontology and provide an idea about the knowledge that needs to be entailed in the ontology.Please use 1. XXXX this format to generate CQ, and do not contain any other content"},
                    {"role": "user", "content": complete_prompt}
                ],
            )
            # response = openai.ChatCompletion.create(
            #     engine="qwen-max", # Specify GPT-4 model
            #     messages=[
            #         {"role": "system", "content": "As an ontology engineer, Provide competency questions focused on the context provided; avoid using narrative questions. competency questions are the questions that outline the scope of an ontology and provide an idea about the knowledge that needs to be entailed in the ontology.Please use 1. XXXX this format to generate CQ, and do not contain any other content"},
            #         {"role": "user", "content": complete_prompt}
            #     ],
            #     temperature=1,
            # )
            question = completion['choices'][0]['message']['content'].strip() # Extract response content
            print("Question:", question)
            questions.append(question)
        except Exception as e:
            print(f"Error generating question for row {row}: {e}")
            questions.append("Error generating question")
    return questions

# Function to read CSV file, generate questions, and save to a single sheet in Excel file
def generate_questions_from_csv(file_path, output_file):
    try:
        df = pd.read_csv(file_path, sep='\t')  # Read CSV file with tab delimiter
    except Exception as e:
        print(f"Error reading CSV file: {e}")  # Handle file reading errors
        return

    # Ensure the output directory exists
    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    # If the file doesn't exist, create a new empty Excel file
    if not os.path.exists(output_file):
        wb = Workbook()
        wb.save(output_file)

    questions = []  # List to store all questions
    try:
        rows = df.values.tolist()  # Convert entire DataFrame to list format
        questions = generate_questions(rows)  # Generate questions for all rows
        df['Question'] = questions  # Add the questions as a new column to the DataFrame

        # Write the entire DataFrame to a single sheet in the Excel file
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name='Questions', index=False)  # Write to a single sheet
        print("Writing to Excel file completed successfully.")

    except Exception as e:
        print(f"Error writing to Excel file: {e}")  # Handle Excel writing errors

# Main process: mount Google Drive, generate questions from CSV, and save results
    
file_path = '../Data/ExtractingTriples/base_ontology.csv' # Define the triple file path
output_questionsFile = 'MyResults/baseontology_qwen.xlsx'#  Define the generated CQs file path
# Ensure the directory exists
# 如果文件不存在，创建一个新的空 Excel 文件
if not os.path.exists(output_questionsFile):
    wb = Workbook()
    wb.save(output_questionsFile)


generate_questions_from_csv(file_path, output_questionsFile)
    
# Play a beep sound after completion
beep()

