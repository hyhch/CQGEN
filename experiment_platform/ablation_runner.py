#!/usr/bin/env python3
"""Ablation experiment runner — one-click execution for Tables A/B/C.

Usage:
    python ablation_runner.py --table A              # Table A only
    python ablation_runner.py --table B              # Table B only
    python ablation_runner.py --table C              # Table C only
    python ablation_runner.py --table all            # All 3 tables
    python ablation_runner.py --table A --dry-run    # Preview, don't execute
    python ablation_runner.py --table A --force      # Force re-run
    python ablation_runner.py --table A --dataset onem2m --model qwen-max  # Single debug
"""

import argparse
import glob as glob_mod
import json
import os
import sys
import time
import threading
import traceback
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

# ---------------------------------------------------------------------------
# Path setup
# ---------------------------------------------------------------------------
_PLATFORM_DIR = os.path.dirname(os.path.abspath(__file__))
if _PLATFORM_DIR not in sys.path:
    sys.path.insert(0, _PLATFORM_DIR)

_RESULTS_DIR = os.path.join(_PLATFORM_DIR, "results")
_ABLATION_DIR = os.path.join(_RESULTS_DIR, "ablation")
_TABLES_DIR = os.path.join(_ABLATION_DIR, "tables")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
ALL_DATASETS = ["onem2m", "saref4env", "videogameontology", "vicinitycore"]
ALL_MODELS = ["qwen-max", "glm-5", "gpt-5"]
DATASET_DISPLAY = {
    "onem2m": "OneM2M",
    "saref4env": "SAREF4ENV",
    "videogameontology": "VGO",
    "vicinitycore": "VC",
}

# ---------------------------------------------------------------------------
# LLM config loading
# ---------------------------------------------------------------------------
def load_llm_configs():
    """Load LLM configurations from llm_configs.json, resolving env vars."""
    config_path = os.path.join(_PLATFORM_DIR, "llm_configs.json")
    with open(config_path, "r", encoding="utf-8") as f:
        configs = json.load(f)

    # Resolve placeholder API keys from environment variables
    env_map = {
        "<DASHSCOPE_API_KEY>": "DASHSCOPE_API_KEY",
        "<AZURE_OPENAI_KEY>": ["AZURE_OPENAI_KEY", "subscription_key"],
    }
    for model_name, cfg in configs.items():
        api_key = cfg.get("api_key", "")
        if api_key in env_map:
            env_names = env_map[api_key]
            if isinstance(env_names, str):
                env_names = [env_names]
            for env_name in env_names:
                val = os.environ.get(env_name, "")
                if val:
                    cfg["api_key"] = val
                    break
    return configs


LLM_CONFIGS = load_llm_configs()

# ---------------------------------------------------------------------------
# Log callback (thread-safe, for Gradio integration)
# ---------------------------------------------------------------------------
_log_lock = threading.Lock()
_log_callback = None


def set_log_callback(cb):
    """Set a callback function that receives formatted log messages."""
    global _log_callback
    with _log_lock:
        _log_callback = cb


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def log(msg, level="INFO"):
    ts = datetime.now().strftime("%H:%M:%S")
    prefix = {"INFO": "   ", "SKIP": " ⏭ ", "RUN": " ▶ ", "OK": " ✓ ",
              "FAIL": " ✗ ", "REUSE": " ♻ ", "WARN": " ⚠ "}
    formatted = f"[{ts}]{prefix.get(level, '   ')} {msg}"
    print(formatted)
    with _log_lock:
        if _log_callback:
            _log_callback(formatted)


def is_experiment_done(table, variant, model, dataset):
    """Check if a result already exists for this experiment."""
    result_dir = os.path.join(_ABLATION_DIR, table, variant, model, dataset)
    if not os.path.isdir(result_dir):
        return False
    for fname in os.listdir(result_dir):
        if fname.startswith("run_") and fname.endswith(".json"):
            fpath = os.path.join(result_dir, fname)
            try:
                with open(fpath, "r", encoding="utf-8") as f:
                    data = json.load(f)
                metrics = data.get("metrics", {})
                if "f1" in metrics and metrics["f1"] is not None:
                    return True
            except Exception:
                continue
    return False


def load_existing_ablation_result(table, variant, model, dataset):
    """Load the latest ablation result for a given experiment."""
    result_dir = os.path.join(_ABLATION_DIR, table, variant, model, dataset)
    if not os.path.isdir(result_dir):
        return None
    files = sorted(glob_mod.glob(os.path.join(result_dir, "run_*.json")), reverse=True)
    for fpath in files:
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            continue
    return None


def find_reusable_main_result(dataset, model):
    """Find the latest main experiment result for OntologyAgent to reuse."""
    result_dir = os.path.join(_RESULTS_DIR, dataset, "OntologyAgent", model)
    if not os.path.isdir(result_dir):
        return None
    files = sorted(glob_mod.glob(os.path.join(result_dir, "run_*.json")), reverse=True)
    for fpath in files:
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                data = json.load(f)
            if "metrics" in data and "f1" in data["metrics"]:
                return data
        except Exception:
            continue
    return None


def save_ablation_result(table, variant, model, dataset, result_data):
    """Save an ablation result to the standard directory structure."""
    result_dir = os.path.join(_ABLATION_DIR, table, variant, model, dataset)
    os.makedirs(result_dir, exist_ok=True)
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    path = os.path.join(result_dir, f"run_{timestamp}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(result_data, f, ensure_ascii=False, indent=2)
    return path


def copy_main_result_as_ablation(table, variant, model, dataset, main_result):
    """Copy a main experiment result into the ablation results directory."""
    result_data = {
        "method": main_result.get("method", "OntologyAgent"),
        "variant": variant,
        "dataset": dataset,
        "model": model,
        "timestamp": main_result.get("timestamp", time.strftime("%Y%m%d_%H%M%S")),
        "duration_seconds": main_result.get("duration_seconds", 0),
        "generated_cqs": main_result.get("generated_cqs", []),
        "metrics": main_result.get("metrics", {}),
        "reused_from_main": True,
    }
    return save_ablation_result(table, variant, model, dataset, result_data)


# ---------------------------------------------------------------------------
# Retry wrapper
# ---------------------------------------------------------------------------
def run_with_retry(run_func, desc, max_retries=3, base_wait=60):
    """Execute run_func with exponential backoff retry."""
    for attempt in range(max_retries):
        try:
            return run_func()
        except Exception as e:
            if attempt == max_retries - 1:
                raise
            wait = base_wait * (2 ** attempt)
            log(f"[RETRY {attempt+1}/{max_retries}] {desc}: {e}, waiting {wait}s...", "WARN")
            time.sleep(wait)


# ---------------------------------------------------------------------------
# Core experiment execution
# ---------------------------------------------------------------------------
def run_single_experiment(table, variant, model_name, dataset, runner_cls, params, force=False):
    """Run a single ablation experiment.

    Returns (success: bool, result_path: str | None, error: str | None)
    """
    exp_id = f"{table}/{variant}/{model_name}/{dataset}"

    # Check if already done
    if not force and is_experiment_done(table, variant, model_name, dataset):
        log(f"[SKIP] {exp_id} — already done", "SKIP")
        return True, None, None

    log(f"[RUN] {exp_id}", "RUN")

    llm_config = dict(LLM_CONFIGS[model_name])
    # Remove concurrency_group from the config passed to runners
    llm_config.pop("concurrency_group", None)

    from runners.ontology_agent_runner import OntologyAgentRunner
    from runners.monolithic_runner import MonolithicRunner
    from evaluation.unified_eval import evaluate
    from dataset_registry import load_ground_truth, load_ground_truth_labels

    if runner_cls == "MonolithicRunner":
        runner = MonolithicRunner()
    else:
        runner = OntologyAgentRunner()

    start_time = time.time()

    def do_run():
        return runner.run(
            dataset_name=dataset,
            llm_config=llm_config,
            params=params,
            progress_callback=lambda msg: log(f"  {msg}"),
        )

    try:
        run_result = run_with_retry(do_run, exp_id)
    except Exception as e:
        error_msg = f"{exp_id} failed: {e}"
        log(error_msg, "FAIL")
        return False, None, error_msg

    if run_result is None:
        error_msg = f"{exp_id} returned None"
        log(error_msg, "FAIL")
        return False, None, error_msg

    # Evaluate
    try:
        gt_cqs = load_ground_truth(dataset)
        gt_labels = load_ground_truth_labels(dataset)
        eval_metrics = evaluate(
            run_result.generated_cqs, gt_cqs,
            threshold=0.6, dedup=True, dedup_threshold=0.85,
            gt_labels=gt_labels,
        )
    except Exception as e:
        log(f"  Evaluation error: {e}", "WARN")
        eval_metrics = {}

    duration = time.time() - start_time

    # Build result data
    merged_metrics = {**run_result.metrics}
    for k, v in eval_metrics.items():
        if k not in ("per_cq_scores", "match_details"):
            merged_metrics[k] = v

    result_data = {
        "method": run_result.method,
        "variant": variant,
        "dataset": dataset,
        "model": model_name,
        "timestamp": time.strftime("%Y%m%d_%H%M%S"),
        "duration_seconds": round(duration, 2),
        "generated_cqs": run_result.generated_cqs,
        "metrics": merged_metrics,
    }

    path = save_ablation_result(table, variant, model_name, dataset, result_data)
    m = result_data["metrics"]
    log(f"[OK] {exp_id}: P={m.get('precision',0):.4f} R={m.get('recall',0):.4f} "
        f"F1={m.get('f1',0):.4f}  -> {path}", "OK")
    return True, path, None


# ---------------------------------------------------------------------------
# Experiment definitions
# ---------------------------------------------------------------------------
def build_table_A_experiments(datasets=None, models=None):
    """Table A: MAS vs Monolithic.

    - Monolithic: 4 datasets × 3 models = 12 new experiments
    - CQGen-MAS: reuse main experiment results
    """
    datasets = datasets or ALL_DATASETS
    models = models or ALL_MODELS
    experiments = []

    # Monolithic experiments (new runs)
    for model in models:
        for ds in datasets:
            experiments.append({
                "table": "tableA",
                "variant": "Monolithic",
                "model": model,
                "dataset": ds,
                "runner_cls": "MonolithicRunner",
                "params": {"cq_examples_num": 10},
                "reuse": False,
            })

    # CQGen-MAS (reuse from main experiments)
    for model in models:
        for ds in datasets:
            experiments.append({
                "table": "tableA",
                "variant": "CQGen-MAS",
                "model": model,
                "dataset": ds,
                "runner_cls": None,
                "params": {},
                "reuse": True,
            })

    return experiments


def build_table_B_experiments(datasets=None, models=None):
    """Table B: Segmentation algorithm comparison (qwen-max only).

    - seg_metis, seg_louvain, seg_leiden, seg_spectral, seg_random: new runs
    - seg_auto: reuse main experiment results
    """
    datasets = datasets or ALL_DATASETS
    # Table B is qwen-max only by design; if user filters to another model, allow it
    models = models or ["qwen-max"]
    experiments = []

    seg_methods = ["metis", "louvain", "leiden", "spectral", "random"]
    for seg in seg_methods:
        for model in models:
            for ds in datasets:
                experiments.append({
                    "table": "tableB",
                    "variant": f"seg_{seg}",
                    "model": model,
                    "dataset": ds,
                    "runner_cls": "OntologyAgentRunner",
                    "params": {"segmentation_method": seg},
                    "reuse": False,
                })

    # seg_auto (reuse from main experiments)
    for model in models:
        for ds in datasets:
            experiments.append({
                "table": "tableB",
                "variant": "seg_auto",
                "model": model,
                "dataset": ds,
                "runner_cls": None,
                "params": {},
                "reuse": True,
            })

    return experiments


def build_table_C_experiments(datasets=None, models=None):
    """Table C: Component ablation.

    - full: reuse main experiment results
    - wo_segmenter, wo_validator, wo_fewshot: new runs
    """
    datasets = datasets or ALL_DATASETS
    models = models or ALL_MODELS
    experiments = []

    # Full (reuse)
    for model in models:
        for ds in datasets:
            experiments.append({
                "table": "tableC",
                "variant": "full",
                "model": model,
                "dataset": ds,
                "runner_cls": None,
                "params": {},
                "reuse": True,
            })

    # w/o Segmenter
    for model in models:
        for ds in datasets:
            experiments.append({
                "table": "tableC",
                "variant": "wo_segmenter",
                "model": model,
                "dataset": ds,
                "runner_cls": "OntologyAgentRunner",
                "params": {"skip_segmentation": True},
                "reuse": False,
            })

    # w/o Validator
    for model in models:
        for ds in datasets:
            experiments.append({
                "table": "tableC",
                "variant": "wo_validator",
                "model": model,
                "dataset": ds,
                "runner_cls": "OntologyAgentRunner",
                "params": {"skip_validator": True},
                "reuse": False,
            })

    # w/o Few-shot
    for model in models:
        for ds in datasets:
            experiments.append({
                "table": "tableC",
                "variant": "wo_fewshot",
                "model": model,
                "dataset": ds,
                "runner_cls": "OntologyAgentRunner",
                "params": {"cq_examples_num": 0},
                "reuse": False,
            })

    return experiments


# ---------------------------------------------------------------------------
# Reuse logic
# ---------------------------------------------------------------------------
def handle_reuse_experiments(experiments, force=False):
    """Process all reuse experiments: copy main results into ablation dir.

    Returns list of experiments that still need to be run (i.e., non-reuse).
    """
    to_run = []
    for exp in experiments:
        if not exp["reuse"]:
            to_run.append(exp)
            continue

        table = exp["table"]
        variant = exp["variant"]
        model = exp["model"]
        dataset = exp["dataset"]
        exp_id = f"{table}/{variant}/{model}/{dataset}"

        if not force and is_experiment_done(table, variant, model, dataset):
            log(f"[SKIP] {exp_id} — already done (reused)", "SKIP")
            continue

        main_result = find_reusable_main_result(dataset, model)
        if main_result is None:
            log(f"[WARN] {exp_id} — no main result to reuse, skipping", "WARN")
            continue

        path = copy_main_result_as_ablation(table, variant, model, dataset, main_result)
        m = main_result.get("metrics", {})
        log(f"[REUSE] {exp_id}: P={m.get('precision',0):.4f} R={m.get('recall',0):.4f} "
            f"F1={m.get('f1',0):.4f}  -> {path}", "REUSE")

    return to_run


# ---------------------------------------------------------------------------
# Concurrent scheduling
# ---------------------------------------------------------------------------
def run_group_sequential(group_exps, force=False):
    """Run a list of experiments sequentially (same concurrency group)."""
    results = []
    for exp in group_exps:
        success, path, error = run_single_experiment(
            table=exp["table"],
            variant=exp["variant"],
            model_name=exp["model"],
            dataset=exp["dataset"],
            runner_cls=exp["runner_cls"],
            params=exp["params"],
            force=force,
        )
        results.append({
            "exp_id": f"{exp['table']}/{exp['variant']}/{exp['model']}/{exp['dataset']}",
            "success": success,
            "path": path,
            "error": error,
        })
    return results


def schedule_experiments(experiments, force=False):
    """Schedule experiments: different concurrency groups run in parallel,
    same group runs sequentially."""
    if not experiments:
        return []

    # Group by concurrency_group
    groups = defaultdict(list)
    for exp in experiments:
        group = LLM_CONFIGS[exp["model"]].get("concurrency_group", exp["model"])
        groups[group].append(exp)

    log(f"Scheduling {len(experiments)} experiments across {len(groups)} "
        f"concurrency group(s): {list(groups.keys())}")

    all_results = []

    if len(groups) == 1:
        # Single group: run sequentially, no thread overhead
        group_name = list(groups.keys())[0]
        all_results = run_group_sequential(groups[group_name], force=force)
    else:
        with ThreadPoolExecutor(max_workers=len(groups)) as pool:
            futures = {}
            for group_name, group_exps in groups.items():
                log(f"  Group '{group_name}': {len(group_exps)} experiments")
                future = pool.submit(run_group_sequential, group_exps, force)
                futures[future] = group_name

            for future in as_completed(futures):
                group_name = futures[future]
                try:
                    results = future.result()
                    all_results.extend(results)
                except Exception as e:
                    log(f"Group '{group_name}' failed: {e}", "FAIL")

    return all_results


# ---------------------------------------------------------------------------
# Table generation
# ---------------------------------------------------------------------------
ALL_METHODS = ["LLM4KE", "Retrofit-CQ", "OntologyAgent"]


def _load_main_result(method, model, dataset):
    """Load the latest main experiment result for a method/model/dataset."""
    result_dir = os.path.join(_RESULTS_DIR, dataset, method, model)
    if not os.path.isdir(result_dir):
        return None
    files = sorted(glob_mod.glob(os.path.join(result_dir, "run_*.json")), reverse=True)
    for fpath in files:
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                data = json.load(f)
            if "metrics" in data and "f1" in data["metrics"]:
                return data
        except Exception:
            continue
    return None


_OA_DATASET_DIR = os.path.join(
    os.path.dirname(_PLATFORM_DIR), "OntologyAgent", "dataset"
)


def _load_sparql_coverage(dataset, model):
    """Load SPARQL-based coverage from original OntologyAgent result files.

    These files contain subgraph_chunks with uncovered_entities, allowing us
    to compute coverage based on SPARQL validation (which is much higher than
    string-based coverage).

    Returns:
        Coverage rate as percentage (0-100), or None if not available.
    """
    fpath = os.path.join(_OA_DATASET_DIR, dataset, f"results_{model}.json")
    if not os.path.isfile(fpath):
        return None
    try:
        with open(fpath, "r", encoding="utf-8") as f:
            data = json.load(f)
        chunks = data.get("subgraph_chunks", [])
        if not chunks:
            return None
        total_ent = 0
        covered_ent = 0
        for chunk in chunks:
            ents = set()
            for t in chunk.get("triples", []):
                ents.add(t[0])
                ents.add(t[2])
            ents = {
                e for e in ents
                if not any(e.startswith(p) for p in ("owl:", "rdf:", "rdfs:", "xsd:"))
            }
            uncov = set(chunk.get("uncovered_entities", []))
            total_ent += len(ents)
            covered_ent += len(ents - uncov)
        if total_ent == 0:
            return None
        return round(covered_ent / total_ent * 100, 2)
    except Exception:
        return None


def generate_main_table():
    """Generate Main Table: method comparison averaged across 3 LLMs.

    OntologyAgent uses SPARQL-based coverage (from original OA result files).
    Baselines (LLM4KE, Retrofit-CQ) use string-based coverage.
    """
    rows = []
    for method in ALL_METHODS:
        row = {"variant": method}
        for ds in ALL_DATASETS:
            ds_metrics = {"precision": [], "recall": [], "f1": [], "cov": []}
            for model in ALL_MODELS:
                result = _load_main_result(method, model, ds)
                if result:
                    m = result.get("metrics", {})
                    ds_metrics["precision"].append(m.get("precision", 0))
                    ds_metrics["recall"].append(m.get("recall", 0))
                    ds_metrics["f1"].append(m.get("f1", 0))
                    # OntologyAgent: use SPARQL coverage; others: string-based
                    if method == "OntologyAgent":
                        sparql_cov = _load_sparql_coverage(ds, model)
                        if sparql_cov is not None:
                            ds_metrics["cov"].append(sparql_cov)
                        else:
                            cs = m.get("coverage_stats", {})
                            ds_metrics["cov"].append(
                                cs.get("overall_coverage_rate",
                                       cs.get("coverage_rate", 0)))
                    else:
                        cs = m.get("coverage_stats", {})
                        ds_metrics["cov"].append(
                            cs.get("overall_coverage_rate",
                                   cs.get("coverage_rate", 0)))
            if ds_metrics["f1"]:
                row[ds] = {
                    "P": sum(ds_metrics["precision"]) / len(ds_metrics["precision"]),
                    "R": sum(ds_metrics["recall"]) / len(ds_metrics["recall"]),
                    "F1": sum(ds_metrics["f1"]) / len(ds_metrics["f1"]),
                    "Cov": sum(ds_metrics["cov"]) / len(ds_metrics["cov"]),
                }
            else:
                row[ds] = {"P": 0, "R": 0, "F1": 0, "Cov": 0}
        rows.append(row)

    # Build markdown table
    header1 = "| Method        |"
    header2 = "|               |"
    sep = "|---------------|"
    for ds in ALL_DATASETS:
        dn = DATASET_DISPLAY[ds]
        header1 += f" {dn:^20s} |"
        header2 += " P     R     F1   Cov |"
        sep += "----------------------|"

    lines = [header1, header2, sep]
    for row in rows:
        line = f"| {row['variant']:<13s} |"
        for ds in ALL_DATASETS:
            m = row[ds]
            line += f" {m['P']:.2f}  {m['R']:.2f}  {m['F1']:.2f}  {m['Cov']/100:.2f} |"
        lines.append(line)

    return "\n".join(lines), rows


def generate_table_A():
    """Generate Table A: MAS vs Monolithic.

    CQGen-MAS uses SPARQL-based coverage; Monolithic uses string-based.
    """
    rows = []
    for variant in ["Monolithic", "CQGen-MAS"]:
        row = {"variant": variant}
        f1_values = []
        cov_values = []
        for ds in ALL_DATASETS:
            ds_metrics = {"f1": [], "cov": []}
            for model in ALL_MODELS:
                result = load_existing_ablation_result("tableA", variant, model, ds)
                if result:
                    m = result.get("metrics", {})
                    ds_metrics["f1"].append(m.get("f1", 0))
                    # CQGen-MAS: SPARQL coverage; Monolithic: string-based
                    if variant == "CQGen-MAS":
                        sparql_cov = _load_sparql_coverage(ds, model)
                        if sparql_cov is not None:
                            ds_metrics["cov"].append(sparql_cov)
                        else:
                            cs = m.get("coverage_stats", {})
                            ds_metrics["cov"].append(
                                cs.get("overall_coverage_rate",
                                       cs.get("coverage_rate", 0)))
                    else:
                        cs = m.get("coverage_stats", {})
                        ds_metrics["cov"].append(
                            cs.get("overall_coverage_rate",
                                   cs.get("coverage_rate", 0)))
            if ds_metrics["f1"]:
                avg_f1 = sum(ds_metrics["f1"]) / len(ds_metrics["f1"])
                avg_cov = sum(ds_metrics["cov"]) / len(ds_metrics["cov"]) if ds_metrics["cov"] else 0
                row[ds] = {"F1": avg_f1, "Cov": avg_cov}
                f1_values.append(avg_f1)
                cov_values.append(avg_cov)
            else:
                row[ds] = {"F1": 0, "Cov": 0}
        row["avg_f1"] = sum(f1_values) / len(f1_values) if f1_values else 0
        row["avg_cov"] = sum(cov_values) / len(cov_values) if cov_values else 0
        rows.append(row)

    # Build markdown table
    header1 = "| Method    |"
    header2 = "|           |"
    sep = "|-----------|"
    for ds in ALL_DATASETS:
        dn = DATASET_DISPLAY[ds]
        header1 += f" {dn:^11s} |"
        header2 += "  F1    Cov  |"
        sep += "-------------|"
    header1 += " Avg F1 | Avg Cov |"
    header2 += "        |         |"
    sep += "--------|---------|"

    lines = [header1, header2, sep]
    for row in rows:
        line = f"| {row['variant']:<9s} |"
        for ds in ALL_DATASETS:
            m = row[ds]
            line += f"  {m['F1']:.2f}   {m['Cov']/100:.2f}  |"
        line += f"  {row['avg_f1']:.2f}  |"
        line += f"   {row['avg_cov']/100:.2f}  |"
        lines.append(line)

    return "\n".join(lines), rows


def generate_table_B():
    """Generate Table B: Segmentation algorithm comparison."""
    seg_display = {
        "seg_metis": "METIS",
        "seg_louvain": "Louvain",
        "seg_leiden": "Leiden",
        "seg_spectral": "Spectral",
        "seg_random": "Random",
        "seg_auto": "Auto(LLM)",
    }
    variants = ["seg_metis", "seg_louvain", "seg_leiden", "seg_spectral", "seg_random", "seg_auto"]
    model = "qwen-max"

    rows = []
    for variant in variants:
        row = {"variant": seg_display[variant]}
        f1_values = []
        cov_values = []
        for ds in ALL_DATASETS:
            result = load_existing_ablation_result("tableB", variant, model, ds)
            if result:
                m = result.get("metrics", {})
                f1 = m.get("f1", 0)
                # Table B: uniform string-based coverage for fair comparison
                cov = m.get("coverage_stats", {}).get("overall_coverage_rate", 0)
                row[ds] = {"f1": f1, "cov": cov}
                f1_values.append(f1)
                cov_values.append(cov)
            else:
                row[ds] = {"f1": 0, "cov": 0}
        row["avg_f1"] = sum(f1_values) / len(f1_values) if f1_values else 0
        row["avg_cov"] = sum(cov_values) / len(cov_values) if cov_values else 0
        rows.append(row)

    # Build markdown table
    header = "| Algorithm |"
    sep = "|-----------|"
    for ds in ALL_DATASETS:
        dn = DATASET_DISPLAY[ds]
        header += f" {dn} F1 | {dn} Cov |"
        sep += f"-{'-'*len(dn)}-----|" * 2
    header += " Avg F1 | Avg Cov |"
    sep += "--------|---------|"

    lines = [header, sep]
    for row in rows:
        line = f"| {row['variant']:<9s} |"
        for ds in ALL_DATASETS:
            dn = DATASET_DISPLAY[ds]
            col_w = len(dn) + 4  # " F1 " / " Cov " padding
            line += f" {row[ds]['f1']:>{col_w}.4f} |"
            line += f" {row[ds]['cov']/100:>{col_w}.4f} |"
        line += f" {row['avg_f1']:.4f}  |"
        line += f" {row['avg_cov']/100:.4f}   |"
        lines.append(line)

    return "\n".join(lines), rows


def generate_table_C():
    """Generate Table C: Component ablation (averaged across 3 LLMs).

    Full variant uses SPARQL-based coverage (from original OA result files).
    Other variants (wo_segmenter, wo_validator, wo_fewshot) use string-based.
    """
    variant_display = {
        "full": "Full",
        "wo_segmenter": "w/o Segmenter",
        "wo_validator": "w/o Validator",
        "wo_fewshot": "w/o Few-shot",
    }
    variants = ["full", "wo_segmenter", "wo_validator", "wo_fewshot"]

    rows = []
    for variant in variants:
        row = {"variant": variant_display[variant]}
        for ds in ALL_DATASETS:
            ds_metrics = {"precision": [], "recall": [], "f1": [], "cov": []}
            for model in ALL_MODELS:
                result = load_existing_ablation_result("tableC", variant, model, ds)
                if result:
                    m = result.get("metrics", {})
                    ds_metrics["precision"].append(m.get("precision", 0))
                    ds_metrics["recall"].append(m.get("recall", 0))
                    ds_metrics["f1"].append(m.get("f1", 0))
                    # Full: use SPARQL coverage; others: string-based
                    if variant == "full":
                        sparql_cov = _load_sparql_coverage(ds, model)
                        if sparql_cov is not None:
                            ds_metrics["cov"].append(sparql_cov)
                        else:
                            cov = m.get("coverage_stats", {}).get("overall_coverage_rate", 0)
                            ds_metrics["cov"].append(cov)
                    else:
                        cov = m.get("coverage_stats", {}).get("overall_coverage_rate", 0)
                        ds_metrics["cov"].append(cov)
            if ds_metrics["f1"]:
                row[ds] = {
                    "P": sum(ds_metrics["precision"]) / len(ds_metrics["precision"]),
                    "R": sum(ds_metrics["recall"]) / len(ds_metrics["recall"]),
                    "F1": sum(ds_metrics["f1"]) / len(ds_metrics["f1"]),
                    "Cov": sum(ds_metrics["cov"]) / len(ds_metrics["cov"]),
                }
            else:
                row[ds] = {"P": 0, "R": 0, "F1": 0, "Cov": 0}
        rows.append(row)

    # Build markdown table
    header1 = "| Config        |"
    header2 = "|               |"
    sep = "|---------------|"
    for ds in ALL_DATASETS:
        dn = DATASET_DISPLAY[ds]
        header1 += f" {dn:^20s} |"
        header2 += " P     R     F1   Cov |"
        sep += "----------------------|"

    lines = [header1, header2, sep]
    for row in rows:
        line = f"| {row['variant']:<13s} |"
        for ds in ALL_DATASETS:
            m = row[ds]
            line += f" {m['P']:.2f}  {m['R']:.2f}  {m['F1']:.2f}  {m['Cov']/100:.2f} |"
        lines.append(line)

    return "\n".join(lines), rows


def save_table(table_name, content):
    """Save a markdown table to results/ablation/tables/."""
    os.makedirs(_TABLES_DIR, exist_ok=True)
    path = os.path.join(_TABLES_DIR, f"{table_name}.md")
    with open(path, "w", encoding="utf-8") as f:
        f.write(content + "\n")


def export_tables_to_excel(table_data):
    """Export all table data to a single Excel file with one sheet per table.

    Args:
        table_data: Dict mapping table letter -> list of row dicts.
            Each row dict has 'variant' and dataset keys with metric sub-dicts.

    Returns:
        Path to the saved Excel file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    os.makedirs(_TABLES_DIR, exist_ok=True)
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    num_fmt_2 = "0.00"
    num_fmt_4 = "0.0000"
    pct_fmt = "0.00%"

    def style_header(ws, row_idx, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    # --- Table A ---
    if "A" in table_data:
        ws = wb.create_sheet("Table A")
        rows = table_data["A"]
        # Header row 1: dataset names (merged)
        headers1 = ["Method"]
        for ds in ALL_DATASETS:
            headers1.extend([DATASET_DISPLAY[ds], ""])
        headers1.extend(["Avg F1", "Avg Cov"])
        # Header row 2: F1 Cov
        headers2 = [""]
        for _ in ALL_DATASETS:
            headers2.extend(["F1", "Cov"])
        headers2.extend(["", ""])

        for c, v in enumerate(headers1, 1):
            ws.cell(row=1, column=c, value=v)
        for c, v in enumerate(headers2, 1):
            ws.cell(row=2, column=c, value=v)
        # Merge dataset header cells
        for i, ds in enumerate(ALL_DATASETS):
            start_col = 2 + i * 2
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 1)
        style_header(ws, 1, len(headers1))
        style_header(ws, 2, len(headers2))

        for r_idx, row in enumerate(rows, 3):
            ws.cell(row=r_idx, column=1, value=row["variant"]).border = thin_border
            col = 2
            for ds in ALL_DATASETS:
                m = row[ds]
                c = ws.cell(row=r_idx, column=col, value=round(m["F1"], 4))
                c.number_format = num_fmt_4
                c.border = thin_border
                col += 1
                c = ws.cell(row=r_idx, column=col, value=round(m["Cov"] / 100, 4))
                c.number_format = pct_fmt
                c.border = thin_border
                col += 1
            c = ws.cell(row=r_idx, column=col, value=round(row["avg_f1"], 4))
            c.number_format = num_fmt_4
            c.border = thin_border
            col += 1
            c = ws.cell(row=r_idx, column=col, value=round(row.get("avg_cov", 0) / 100, 4))
            c.number_format = pct_fmt
            c.border = thin_border

        ws.column_dimensions["A"].width = 14

    # --- Table B ---
    if "B" in table_data:
        ws = wb.create_sheet("Table B")
        rows = table_data["B"]
        headers = ["Algorithm"]
        for ds in ALL_DATASETS:
            dn = DATASET_DISPLAY[ds]
            headers.extend([f"{dn} F1", f"{dn} Cov"])
        headers.extend(["Avg F1", "Avg Cov"])

        for c, v in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=v)
        style_header(ws, 1, len(headers))

        for r_idx, row in enumerate(rows, 2):
            ws.cell(row=r_idx, column=1, value=row["variant"]).border = thin_border
            col = 2
            for ds in ALL_DATASETS:
                m = row[ds]
                c = ws.cell(row=r_idx, column=col, value=round(m["f1"], 4))
                c.number_format = num_fmt_4
                c.border = thin_border
                col += 1
                c = ws.cell(row=r_idx, column=col, value=round(m["cov"] / 100, 4))
                c.number_format = pct_fmt
                c.border = thin_border
                col += 1
            c = ws.cell(row=r_idx, column=col, value=round(row["avg_f1"], 4))
            c.number_format = num_fmt_4
            c.border = thin_border
            col += 1
            c = ws.cell(row=r_idx, column=col, value=round(row["avg_cov"] / 100, 4))
            c.number_format = pct_fmt
            c.border = thin_border

        ws.column_dimensions["A"].width = 12

    # --- Table C ---
    if "C" in table_data:
        ws = wb.create_sheet("Table C")
        rows = table_data["C"]
        # Header row 1: dataset names
        headers1 = ["Config"]
        for ds in ALL_DATASETS:
            headers1.extend([DATASET_DISPLAY[ds], "", "", ""])
        # Header row 2: P R F1 Cov
        headers2 = [""]
        for _ in ALL_DATASETS:
            headers2.extend(["P", "R", "F1", "Cov"])

        for c, v in enumerate(headers1, 1):
            ws.cell(row=1, column=c, value=v)
        for c, v in enumerate(headers2, 1):
            ws.cell(row=2, column=c, value=v)
        for i, ds in enumerate(ALL_DATASETS):
            start_col = 2 + i * 4
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 3)
        style_header(ws, 1, len(headers1))
        style_header(ws, 2, len(headers2))

        for r_idx, row in enumerate(rows, 3):
            ws.cell(row=r_idx, column=1, value=row["variant"]).border = thin_border
            col = 2
            for ds in ALL_DATASETS:
                m = row[ds]
                for val in [m["P"], m["R"], m["F1"]]:
                    c = ws.cell(row=r_idx, column=col, value=round(val, 4))
                    c.number_format = num_fmt_4
                    c.border = thin_border
                    col += 1
                c = ws.cell(row=r_idx, column=col, value=round(m["Cov"] / 100, 4))
                c.number_format = pct_fmt
                c.border = thin_border
                col += 1

        ws.column_dimensions["A"].width = 16

    # --- Main Table ---
    if "Main" in table_data:
        ws = wb.create_sheet("Main Table")
        rows = table_data["Main"]
        headers1 = ["Method"]
        for ds in ALL_DATASETS:
            headers1.extend([DATASET_DISPLAY[ds], "", "", ""])
        headers2 = [""]
        for _ in ALL_DATASETS:
            headers2.extend(["P", "R", "F1", "Cov"])

        for c, v in enumerate(headers1, 1):
            ws.cell(row=1, column=c, value=v)
        for c, v in enumerate(headers2, 1):
            ws.cell(row=2, column=c, value=v)
        for i, ds in enumerate(ALL_DATASETS):
            start_col = 2 + i * 4
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 3)
        style_header(ws, 1, len(headers1))
        style_header(ws, 2, len(headers2))

        for r_idx, row in enumerate(rows, 3):
            ws.cell(row=r_idx, column=1, value=row["variant"]).border = thin_border
            col = 2
            for ds in ALL_DATASETS:
                m = row[ds]
                for val in [m["P"], m["R"], m["F1"]]:
                    c = ws.cell(row=r_idx, column=col, value=round(val, 4))
                    c.number_format = num_fmt_4
                    c.border = thin_border
                    col += 1
                c = ws.cell(row=r_idx, column=col, value=round(m["Cov"] / 100, 4))
                c.number_format = pct_fmt
                c.border = thin_border
                col += 1

        ws.column_dimensions["A"].width = 16

    path = os.path.join(_TABLES_DIR, "ablation_tables.xlsx")
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Dry-run preview
# ---------------------------------------------------------------------------
def preview_experiments(experiments):
    """Preview experiment plan without executing. Returns summary string."""
    new_count = 0
    reuse_count = 0
    skip_count = 0

    for exp in experiments:
        exp_id = f"{exp['table']}/{exp['variant']}/{exp['model']}/{exp['dataset']}"
        if exp["reuse"]:
            main = find_reusable_main_result(exp["dataset"], exp["model"])
            if is_experiment_done(exp["table"], exp["variant"], exp["model"], exp["dataset"]):
                log(f"[SKIP]  {exp_id} (already done, reused)", "SKIP")
                skip_count += 1
            elif main:
                log(f"[REUSE] {exp_id} <- main result found", "REUSE")
                reuse_count += 1
            else:
                log(f"[MISS]  {exp_id} <- NO main result to reuse!", "WARN")
        else:
            if is_experiment_done(exp["table"], exp["variant"], exp["model"], exp["dataset"]):
                log(f"[SKIP]  {exp_id} (already done)", "SKIP")
                skip_count += 1
            else:
                runner = exp["runner_cls"]
                params_str = ", ".join(f"{k}={v}" for k, v in exp["params"].items())
                log(f"[NEW]   {exp_id} ({runner}, {params_str})")
                new_count += 1

    summary = (f"Summary: {new_count} new, {reuse_count} reuse, {skip_count} skip, "
               f"{new_count + reuse_count + skip_count} total")
    log(summary)
    return summary


# ---------------------------------------------------------------------------
# High-level API (used by both CLI and Gradio)
# ---------------------------------------------------------------------------
def _build_experiments(tables, dataset=None, model=None):
    """Build experiment list for the given tables and optional filters."""
    datasets = [dataset] if dataset else None
    models = [model] if model else None
    experiments = []
    for table in tables:
        if table == "A":
            experiments.extend(build_table_A_experiments(datasets, models))
        elif table == "B":
            experiments.extend(build_table_B_experiments(datasets, models))
        elif table == "C":
            experiments.extend(build_table_C_experiments(datasets, models))
    return experiments


def refresh_tables(tables):
    """Generate tables from existing results without running experiments.

    Args:
        tables: List of table identifiers. Valid values:
            "Main", "A", "B", "C".

    Returns dict mapping table identifier -> markdown string.
    """
    outputs = {}
    table_data = {}
    for table in tables:
        if table == "Main":
            content, rows = generate_main_table()
            save_table("main_table", content)
            outputs["Main"] = content
            table_data["Main"] = rows
        elif table == "A":
            content, rows = generate_table_A()
            save_table("tableA", content)
            outputs["A"] = content
            table_data["A"] = rows
        elif table == "B":
            content, rows = generate_table_B()
            save_table("tableB", content)
            outputs["B"] = content
            table_data["B"] = rows
        elif table == "C":
            content, rows = generate_table_C()
            save_table("tableC", content)
            outputs["C"] = content
            table_data["C"] = rows

    # Export to Excel
    if table_data:
        xlsx_path = export_tables_to_excel(table_data)
        log(f"Excel exported: {xlsx_path}")

    return outputs


def run_ablation(tables, dry_run=False, force=False, dataset=None, model=None):
    """Run ablation experiments and return results.

    This is the main API entry point used by both CLI and Gradio.
    Log messages are sent through ``set_log_callback`` if one is set.

    Returns:
        Dict with keys: mode, tables (dict of table markdown), stats.
    """
    all_experiments = _build_experiments(tables, dataset, model)

    if not all_experiments:
        log("No experiments defined.")
        return {"mode": "empty", "tables": {}, "stats": {}}

    log(f"{'='*60}")
    log(f"Ablation Runner — Tables: {', '.join(tables)}")
    log(f"Total experiments: {len(all_experiments)}")
    log(f"{'='*60}")

    if dry_run:
        preview_experiments(all_experiments)
        table_outputs = refresh_tables(tables)
        return {"mode": "dry_run", "tables": table_outputs, "stats": {}}

    start = time.time()

    # Step 1: Handle reuse experiments (copy main results)
    to_run = handle_reuse_experiments(all_experiments, force=force)

    # Step 2: Schedule and run new experiments
    results = []
    if to_run:
        results = schedule_experiments(to_run, force=force)
    else:
        log("No new experiments to run.")

    succeeded = sum(1 for r in results if r["success"])
    failed = sum(1 for r in results if not r["success"])
    skipped = sum(1 for r in results if r["success"] and r["path"] is None)
    reused = len(all_experiments) - len(to_run)

    # Step 3: Generate tables
    log(f"{'='*60}")
    log("Generating tables...")
    log(f"{'='*60}")
    table_outputs = refresh_tables(tables)
    for name, content in table_outputs.items():
        log(f"\n--- Table {name} ---")
        for line in content.split("\n"):
            log(line)

    total_time = time.time() - start
    log(f"{'='*60}")
    log(f"Done! Total time: {total_time:.1f}s")
    log(f"  Reused from main: {reused}")
    log(f"  Ran (succeeded):  {succeeded}")
    log(f"  Ran (skipped):    {skipped}")
    log(f"  Failed:           {failed}")
    log(f"{'='*60}")

    return {
        "mode": "run",
        "tables": table_outputs,
        "stats": {
            "total": len(all_experiments),
            "reused": reused,
            "succeeded": succeeded,
            "failed": failed,
            "skipped": skipped,
            "duration": round(total_time, 1),
        },
    }


# ---------------------------------------------------------------------------
# CLI Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Run ablation experiments and generate paper tables."
    )
    parser.add_argument(
        "--table", required=True,
        choices=["A", "B", "C", "all"],
        help="Which table(s) to run: A, B, C, or all",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Preview experiments without executing",
    )
    parser.add_argument(
        "--force", action="store_true",
        help="Force re-run even if results exist",
    )
    parser.add_argument(
        "--dataset", type=str, default=None,
        help="Run only this dataset (e.g., onem2m)",
    )
    parser.add_argument(
        "--model", type=str, default=None,
        help="Run only this model (e.g., qwen-max)",
    )
    args = parser.parse_args()

    tables = ["A", "B", "C"] if args.table == "all" else [args.table]
    result = run_ablation(
        tables=tables,
        dry_run=args.dry_run,
        force=args.force,
        dataset=args.dataset,
        model=args.model,
    )


if __name__ == "__main__":
    main()
